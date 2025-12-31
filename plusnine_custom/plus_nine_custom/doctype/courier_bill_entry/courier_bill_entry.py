# Copyright (c) 2025, sanprasoftwares@gmail.com and contributors
# For license information, please see license.txt

import frappe
import io
from frappe.model.document import Document
from openpyxl import load_workbook  


class CourierBillEntry(Document):  
	@frappe.whitelist()  
	def attch_excel_file(self):
		if self.attach:
			file_bytes = frappe.get_doc("File", {"file_url": self.attach}).get_content()
			workbook = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
			sheet = workbook.active

			self.set("items", [])  

			selected_transporter = self.transporter

			for row in range(2, sheet.max_row + 1):
				transporter_name = sheet.cell(row=row, column=1).value
				docket_number = sheet.cell(row=row, column=2).value
				date = sheet.cell(row=row, column=3).value
				parcel_type = sheet.cell(row=row, column=4).value
				delivered_to = sheet.cell(row=row, column=5).value
				weight = sheet.cell(row=row, column=6).value
				amount_charged = sheet.cell(row=row, column=7).value  
				soname = sheet.cell(row=row, column=8).value

				if not selected_transporter or transporter_name == selected_transporter:
					self.append("items", {
						"transporter_name": transporter_name,
						"docket_number": docket_number,
						"date": date, 
						"parcel_type": parcel_type,
						"delivered_to": delivered_to,
						"weight": weight,
						"amount_charged": amount_charged,
						"link_sales_invoice": soname,
						"reconciliation_status": ""
					})

	@frappe.whitelist()
	def match_sales_invoices(self):
		matched = 0

		for row in self.get("items"):
			si = frappe.db.get_value(
				"Sales Invoice",
				{
					"lr_no": row.docket_number,
					"transporter_name": row.transporter_name
				},
				["name", "grand_total", "total_net_weight", "shipping_address_name"]  
			)
			if si:
				name, total, total_net_weight, shipping_address  = si
				row.link_sales_invoice = name

				if total_net_weight:
					row.total_weight = total_net_weight
				
				if shipping_address:
					city = frappe.db.get_value("Address", shipping_address, "city")
					if city:
						row.city = city   

				row.reconciliation_status = "Matched"
				matched += 1
			else:
				row.reconciliation_status = "Unmatched"

		return f"{matched} entries matched."

